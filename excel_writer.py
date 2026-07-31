"""Write call updates and audit logs back to Excel."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import LectureRecord
from utils import isoformat_or_empty


CALL_LOG_HEADERS = [
    "Timestamp",
    "Teacher ID",
    "Teacher Name",
    "Phone Number",
    "Department",
    "Subject",
    "Lecture Date",
    "Lecture Time",
    "Room",
    "Attempt Number",
    "Call SID",
    "Call Status",
    "Retry Count",
    "Next Call Time",
    "Teacher Response",
    "Delay Minutes",
    "Call Duration",
    "Conversation Transcript",
    "Reason",
    "Decision JSON",
]


class ExcelWriter:
    """Persist the latest call state and immutable call history."""

    def __init__(self, workbook_path: Path, sheet_name: str, call_logs_sheet_name: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.call_logs_sheet_name = call_logs_sheet_name

    def update_lecture_row(self, lecture: LectureRecord) -> None:
        """Update the source workbook row with the latest mutable fields."""

        if not self.workbook_path.exists():
            return

        workbook = load_workbook(self.workbook_path)
        worksheet = self._get_sheet(workbook, self.sheet_name)
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
        header_map = {str(cell.value).strip(): index + 1 for index, cell in enumerate(header_row) if cell.value is not None}

        target_row = lecture.source_row or self._find_row(worksheet, lecture)
        if target_row is None:
            workbook.close()
            return

        self._set_cell(worksheet, target_row, header_map, "Call Status", lecture.call_status)
        self._set_cell(worksheet, target_row, header_map, "Retry Count", lecture.retry_count)
        self._set_cell(worksheet, target_row, header_map, "Next Call Time", isoformat_or_empty(lecture.next_call_time))
        self._set_cell(worksheet, target_row, header_map, "Teacher Response", lecture.teacher_response)
        self._set_cell(worksheet, target_row, header_map, "Delay Minutes", lecture.delay_minutes)
        self._set_cell(worksheet, target_row, header_map, "Call Attempts", lecture.call_attempts)
        self._set_cell(worksheet, target_row, header_map, "Last Call Time", isoformat_or_empty(lecture.last_call_time))
        self._set_cell(worksheet, target_row, header_map, "Conversation Transcript", lecture.conversation_transcript)
        self._set_cell(worksheet, target_row, header_map, "Reason", lecture.reason)
        workbook.save(self.workbook_path)
        workbook.close()

    def append_call_log(self, row_data: dict[str, Any]) -> None:
        """Append a new row to the Call Logs worksheet without overwriting history."""

        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path)
        else:
            workbook = self._create_workbook()

        worksheet = self._get_sheet(workbook, self.call_logs_sheet_name, create=True)
        if worksheet.max_row == 1 and all(cell.value is None for cell in worksheet[1]):
            worksheet.delete_rows(1, 1)
            worksheet.append(CALL_LOG_HEADERS)
        elif worksheet.max_row == 0:
            worksheet.append(CALL_LOG_HEADERS)

        worksheet.append([
            row_data.get("Timestamp", ""),
            row_data.get("Teacher ID", ""),
            row_data.get("Teacher Name", ""),
            row_data.get("Phone Number", ""),
            row_data.get("Department", ""),
            row_data.get("Subject", ""),
            row_data.get("Lecture Date", ""),
            row_data.get("Lecture Time", ""),
            row_data.get("Room", ""),
            row_data.get("Attempt Number", ""),
            row_data.get("Call SID", ""),
            row_data.get("Call Status", ""),
            row_data.get("Retry Count", ""),
            row_data.get("Next Call Time", ""),
            row_data.get("Teacher Response", ""),
            row_data.get("Delay Minutes", ""),
            row_data.get("Call Duration", ""),
            row_data.get("Conversation Transcript", ""),
            row_data.get("Reason", ""),
            row_data.get("Decision JSON", ""),
        ])
        workbook.save(self.workbook_path)
        workbook.close()

    def _create_workbook(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name
        sheet.append([
            "Teacher ID",
            "Teacher Name",
            "Phone Number",
            "Department",
            "Subject",
            "Lecture Date",
            "Lecture Time",
            "Room",
            "Call Status",
            "Retry Count",
            "Next Call Time",
            "Teacher Response",
            "Delay Minutes",
            "Call Attempts",
            "Last Call Time",
            "Conversation Transcript",
            "Reason",
        ])
        workbook.create_sheet(self.call_logs_sheet_name)
        return workbook

    def _get_sheet(self, workbook, sheet_name: str, create: bool = False) -> Worksheet:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        if not create:
            return workbook[workbook.sheetnames[0]]
        return workbook.create_sheet(sheet_name)

    def _find_row(self, worksheet: Worksheet, lecture: LectureRecord) -> int | None:
        """Best-effort lookup when the source row number is unavailable."""

        for row_index in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_index, column=column_index).value for column_index in range(1, worksheet.max_column + 1)]
            text = [str(value).strip() if value is not None else "" for value in values]
            if (
                len(text) >= 7
                and text[0] == lecture.teacher_id
                and text[1] == lecture.teacher_name
                and text[4] == lecture.subject
                and text[5] == lecture.lecture_date.isoformat()
            ):
                return row_index
        return None

    def _set_cell(self, worksheet: Worksheet, row: int, header_map: dict[str, int], header: str, value: Any) -> None:
        column_index = header_map.get(header)
        if column_index is not None:
            worksheet.cell(row=row, column=column_index, value=value)
