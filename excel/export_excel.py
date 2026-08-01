"""Excel schedule export."""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from config import get_settings
from models import Lecture

try:
    import pythoncom
    from win32com import client as win32_client
except ImportError:  # pragma: no cover - Windows-only fallback dependency
    pythoncom = None
    win32_client = None

logger = logging.getLogger(__name__)

EXPORT_COLUMNS = [
    "Teacher ID",
    "Teacher Name",
    "Phone Number",
    "Department",
    "Subject",
    "Lecture Date",
    "Lecture Time",
    "Room",
    "Confirmation Status",
    "Retry Count",
    "Next Retry Time",
    "Last Call Time",
    "Teacher Response",
    "Delay Minutes",
    "Reason",
    "Transcript",
    "Conversation Finished",
]


class ExcelExporter:
    """Export lecture schedule with confirmation data to Excel."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.settings = get_settings()

    def export_file(self, file_path: str | Path | None = None) -> Path:
        """Export all lectures to Excel and return the file path."""
        path = Path(file_path) if file_path else Path(self.settings.excel_file_path)
        if not path.is_absolute():
            path = self.settings.data_dir / path.name

        lectures = (
            self.db.query(Lecture)
            .options(joinedload(Lecture.teacher))
            .order_by(Lecture.lecture_date, Lecture.lecture_time)
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Lecture Schedule"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, lecture in enumerate(lectures, start=2):
            teacher = lecture.teacher
            row_data = [
                teacher.teacher_id if teacher else "",
                teacher.name if teacher else "",
                teacher.phone_number if teacher else "",
                teacher.department if teacher else "",
                lecture.subject,
                lecture.lecture_date.strftime("%Y-%m-%d") if lecture.lecture_date else "",
                lecture.lecture_time,
                lecture.room,
                lecture.confirmation_status,
                lecture.retry_count,
                lecture.next_retry_time.isoformat() if lecture.next_retry_time else "",
                lecture.last_call_time.isoformat() if lecture.last_call_time else "",
                lecture.teacher_response or "",
                lecture.delay_minutes,
                lecture.reason or "",
                lecture.transcript or "",
                "Yes" if lecture.conversation_finished else "No",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.save(path)
        except PermissionError as exc:
            logger.warning("Primary Excel save failed for %s: %s", path, exc)
            self._save_with_excel_com(path, lectures)
        logger.info("Exported %d lectures to %s", len(lectures), path)
        return path

    def _save_with_excel_com(self, path: Path, lectures: list[Lecture]) -> None:
        """Update an already-open workbook through Excel automation on Windows."""
        if pythoncom is None or win32_client is None:
            raise PermissionError(f"Unable to save locked workbook: {path}")

        excel = None
        workbook = None
        launched_excel = False
        opened_workbook = False

        try:
            pythoncom.CoInitialize()
            try:
                excel = win32_client.GetActiveObject("Excel.Application")
            except Exception:
                excel = win32_client.Dispatch("Excel.Application")
                launched_excel = True

            workbook = self._find_open_workbook(excel, path)
            if workbook is None:
                workbook = excel.Workbooks.Open(str(path))
                opened_workbook = True

            worksheet = workbook.Worksheets("Lecture Schedule")
            self._write_rows_via_com(worksheet, lectures)
            workbook.Save()
        except Exception as exc:
            logger.error("Excel COM fallback failed for %s: %s", path, exc)
            raise
        finally:
            if workbook is not None and opened_workbook:
                workbook.Close(SaveChanges=True)
            if excel is not None and launched_excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    @staticmethod
    def _find_open_workbook(excel: Any, path: Path):
        target = str(path.resolve()).lower()
        try:
            for workbook in excel.Workbooks:
                try:
                    if str(Path(workbook.FullName).resolve()).lower() == target:
                        return workbook
                except Exception:
                    continue
        except Exception:
            return None
        return None

    def _write_rows_via_com(self, worksheet: Any, lectures: list[Lecture]) -> None:
        rows = []
        for lecture in lectures:
            teacher = lecture.teacher
            rows.append([
                teacher.teacher_id if teacher else "",
                teacher.name if teacher else "",
                teacher.phone_number if teacher else "",
                teacher.department if teacher else "",
                lecture.subject,
                lecture.lecture_date.strftime("%Y-%m-%d") if lecture.lecture_date else "",
                lecture.lecture_time,
                lecture.room,
                lecture.confirmation_status,
                lecture.retry_count,
                lecture.next_retry_time.isoformat() if lecture.next_retry_time else "",
                lecture.last_call_time.isoformat() if lecture.last_call_time else "",
                lecture.teacher_response or "",
                lecture.delay_minutes,
                lecture.reason or "",
                lecture.transcript or "",
                "Yes" if lecture.conversation_finished else "No",
            ])

        for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
            worksheet.Cells(1, col_idx).Value = header

        used_rows = int(getattr(worksheet.UsedRange, "Rows").Count or 1)
        last_data_row = max(2, used_rows)
        if last_data_row > len(rows) + 1:
            worksheet.Range(f"A{len(rows) + 2}:Q{last_data_row}").ClearContents()

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.Cells(row_idx, col_idx).Value = value
