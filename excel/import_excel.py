"""Excel schedule import."""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from config import get_settings
from services.lecture_service import LectureService
from services.teacher_service import TeacherService

logger = logging.getLogger(__name__)

EXPECTED_COLUMNS = [
    "Teacher ID",
    "Teacher Name",
    "Phone Number",
    "Department",
    "Subject",
    "Lecture Date",
    "Lecture Time",
    "Room",
]


class ExcelImporter:
    """Import lecture schedule from Excel into the database."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.settings = get_settings()
        self.teacher_service = TeacherService(db)
        self.lecture_service = LectureService(db)

    def import_file(self, file_path: str | Path | None = None) -> dict:
        """Import Excel file and return summary stats."""
        path = Path(file_path) if file_path else Path(self.settings.excel_file_path)
        if not path.is_absolute():
            path = self.settings.base_dir / path

        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Excel file is empty")

        headers = [str(h).strip() if h else "" for h in rows[0]]
        self._validate_headers(headers)

        col_map = {h: i for i, h in enumerate(headers)}
        imported = 0
        errors: list[str] = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                if not row or not row[col_map["Teacher ID"]]:
                    continue
                self._import_row(row, col_map)
                imported += 1
            except Exception as exc:
                errors.append(f"Row {row_num}: {exc}")
                logger.warning("Import error row %d: %s", row_num, exc)

        self.db.commit()
        logger.info("Imported %d lectures from %s", imported, path)
        return {"imported": imported, "errors": errors, "file": str(path)}

    def _import_row(self, row: tuple, col_map: dict[str, int]) -> None:
        def val(key: str) -> str:
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip() if row[idx] is not None else ""

        teacher = self.teacher_service.create_or_update(
            teacher_id=val("Teacher ID"),
            name=val("Teacher Name"),
            phone_number=val("Phone Number"),
            department=val("Department"),
        )

        date_str = val("Lecture Date")
        lecture_date = self._parse_date(date_str)

        self.lecture_service.create_or_update(
            teacher=teacher,
            subject=val("Subject"),
            lecture_date=lecture_date,
            lecture_time=val("Lecture Time"),
            room=val("Room"),
        )

    @staticmethod
    def _parse_date(date_str: str) -> datetime:
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        if isinstance(date_str, datetime):
            return date_str
        raise ValueError(f"Unrecognized date format: {date_str}")

    @staticmethod
    def _validate_headers(headers: list[str]) -> None:
        missing = [c for c in EXPECTED_COLUMNS if c not in headers]
        if missing:
            raise ValueError(f"Missing required columns: {missing}")
