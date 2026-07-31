"""Generate sample Excel schedule file."""

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

COLUMNS = [
    "Teacher ID",
    "Teacher Name",
    "Phone Number",
    "Department",
    "Subject",
    "Lecture Date",
    "Lecture Time",
    "Room",
    "Confirmation Status",
    "Retry Count",
    "Next Retry Time",
    "Last Call Time",
    "Teacher Response",
    "Delay Minutes",
    "Reason",
    "Transcript",
    "Conversation Finished",
]

SAMPLE_DATA = [
    ("T001", "Professor Amit Sharma", "+91-9876543210", "Computer Science",
     "Database Management Systems", "10:00 AM", "Lab-301"),
    ("T002", "Dr. Priya Patel", "+91-9876543211", "Electronics",
     "Digital Signal Processing", "11:00 AM", "Room-205"),
    ("T003", "Professor Rajesh Kumar", "+91-9876543212", "Mechanical",
     "Thermodynamics", "09:00 AM", "Room-101"),
    ("T004", "Dr. Sneha Reddy", "+91-9876543213", "Computer Science",
     "Machine Learning", "02:00 PM", "Lab-302"),
    ("T005", "Professor Vikram Singh", "+91-9876543214", "Civil",
     "Structural Analysis", "03:00 PM", "Room-110"),
]


def create_sample_excel(output_path: Path) -> None:
    tomorrow = date.today() + timedelta(days=1)
    date_str = tomorrow.strftime("%Y-%m-%d")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecture Schedule"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (tid, name, phone, dept, subject, time, room) in enumerate(SAMPLE_DATA, start=2):
        row_data = [
            tid, name, phone, dept, subject, date_str, time, room,
            "Pending", 0, "", "", "", 0, "", "", "No",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Sample Excel created at {output_path}")


if __name__ == "__main__":
    base = Path(__file__).resolve().parent.parent
    create_sample_excel(base / "data" / "lecture_schedule.xlsx")
